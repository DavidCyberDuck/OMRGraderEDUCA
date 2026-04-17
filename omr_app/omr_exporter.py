"""
Excel Exporter — results to .xlsx with 4 sheets and charts.
Column order: Folio → Grado → Grupo → Score...
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import datetime

C_HDR_BG = "2C3E50"
C_HDR_FG = "FFFFFF"
C_ACCENT = "3498DB"
C_GREEN  = "D5F5E3"
C_RED    = "FADBD8"
C_ALT    = "F8F9FA"
C_BORDER = "BDC3C7"


def _hdr(cell, bg=C_HDR_BG, fg=C_HDR_FG):
    cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = _border()


def _border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w


def export_to_excel(grade_results, answer_key, exam_name, output_path):
    wb = Workbook()
    _summary(wb, grade_results, exam_name)
    _detail(wb, grade_results, answer_key)
    _sk_sheet(wb, grade_results)
    _charts(wb, grade_results)
    _clave_sheet(wb, answer_key, exam_name)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(output_path)
    return output_path


def _summary(wb, results, exam_name):
    ws = wb.create_sheet("Resumen")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"].value = f"Resultados — {exam_name}"
    ws["A1"].font  = Font(bold=True, size=14, color=C_HDR_FG, name="Arial")
    ws["A1"].fill  = PatternFill("solid", fgColor=C_HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font  = Font(italic=True, size=9, color="888888", name="Arial")
    ws["A2"].alignment = Alignment(horizontal="right")

    headers = ["Folio","Grado","Grupo","Puntaje","Total",
               "Porcentaje (%)","Calificación","Prom. Autoconoc.",
               "Confianza (%)","Estado"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(row=4, column=col, value=h))
    ws.row_dimensions[4].height = 32

    grade_map = [(90,"A"),(80,"B"),(70,"C"),(60,"D"),(0,"F")]

    for r, gr in enumerate(results, 5):
        bg     = C_ALT if r % 2 == 0 else "FFFFFF"
        letter = next(l for pct, l in grade_map if gr.percentage >= pct)
        row_data = [
            gr.folio, gr.grado or "?", gr.grupo or "?",
            gr.score, gr.total, gr.percentage, letter,
            gr.sk_average if gr.sk_average is not None else "N/A",
            round(gr.confidence * 100, 0),
            "⚠ Error" if gr.error else "OK",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()

        # Color-code percentage column (col 6)
        pct_cell = ws.cell(row=r, column=6)
        pct_cell.fill = PatternFill("solid",
                        fgColor=C_GREEN if gr.percentage >= 70 else C_RED)

    last = 4 + len(results)
    sr   = last + 2
    ws.cell(row=sr, column=1, value="Estadísticas").font = Font(bold=True, name="Arial")
    for i, (lbl, fml) in enumerate([
        ("Promedio",         f"=AVERAGE(F5:F{last})"),
        ("Máximo",           f"=MAX(F5:F{last})"),
        ("Mínimo",           f"=MIN(F5:F{last})"),
        ("Aprobados (≥70%)", f'=COUNTIF(F5:F{last},">=70")'),
    ]):
        ws.cell(row=sr+1+i, column=1, value=lbl).font = Font(bold=True, name="Arial", size=10)
        ws.cell(row=sr+1+i, column=2, value=fml).font = Font(name="Arial", size=10)

    for col, w in enumerate([8,8,8,10,8,16,14,18,14,10], 1):
        _cw(ws, col, w)

    ws.conditional_formatting.add(f"F5:F{last}", ColorScaleRule(
        start_type="num", start_value=0,   start_color="E74C3C",
        mid_type="num",   mid_value=70,    mid_color="F39C12",
        end_type="num",   end_value=100,   end_color="2ECC71",
    ))


def _detail(wb, results, answer_key):
    ws  = wb.create_sheet("Detalle Preguntas")
    ws.sheet_view.showGridLines = False
    n_q = len(answer_key)

    ws.merge_cells(f"A1:{get_column_letter(n_q+5)}1")
    ws["A1"].value = "Detalle de Respuestas — Sección 1"
    ws["A1"].font  = Font(bold=True, size=13, color=C_HDR_FG, name="Arial")
    ws["A1"].fill  = PatternFill("solid", fgColor=C_HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, val in enumerate(["Clave","Folio","Grado","Grupo"] +
                               answer_key, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = Font(bold=True,
                         color=C_HDR_FG if col > 4 else "000000",
                         name="Arial", size=9)
        if col > 4:
            cell.fill = PatternFill("solid", fgColor=C_ACCENT)
        cell.alignment = Alignment(horizontal="center")

    for col, val in enumerate(["","Folio","Grado","Grupo"] +
                               [f"P{q}" for q in range(1, n_q+1)], 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.font = Font(bold=True, name="Arial", size=8)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="ECF0F1")

    for r, gr in enumerate(results, 4):
        ws.cell(row=r, column=1, value=gr.page_num)
        ws.cell(row=r, column=2, value=gr.folio)
        ws.cell(row=r, column=3, value=gr.grado or "?")
        ws.cell(row=r, column=4, value=gr.grupo or "?")
        for q, (ans, ok) in enumerate(zip(gr.mc_answers, gr.mc_correct)):
            cell = ws.cell(row=r, column=q+5, value=ans or "-")
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor=C_GREEN if ok else C_RED)

    last = 3 + len(results)
    acc  = last + 2
    ws.cell(row=acc, column=1, value="Aciertos %").font = Font(bold=True, name="Arial", size=9)
    for q in range(n_q):
        correct = sum(1 for gr in results
                      if q < len(gr.mc_correct) and gr.mc_correct[q])
        pct  = round(correct / max(len(results), 1) * 100, 1)
        cell = ws.cell(row=acc, column=q+5, value=pct)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid",
                    fgColor=C_GREEN if pct >= 80 else (C_RED if pct < 50 else "FFFFFF"))

    for col in range(1, n_q+6):
        _cw(ws, col, 5.5)
    for col, w in enumerate([6,8,7,7], 1):
        _cw(ws, col, w)


def _sk_sheet(wb, results):
    ws = wb.create_sheet("Autoconocimiento")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    ws["A1"].value = "Sección 2 — Autoconocimiento"
    ws["A1"].font  = Font(bold=True, size=13, color=C_HDR_FG, name="Arial")
    ws["A1"].fill  = PatternFill("solid", fgColor=C_HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Folio","Grado","Grupo"] + \
              [f"SK{i}" for i in range(1,11)] + ["Promedio"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=col, value=h), bg=C_ACCENT)

    for r, gr in enumerate(results, 3):
        bg = C_ALT if r % 2 == 0 else "FFFFFF"
        for col, val in enumerate([gr.folio, gr.grado or "?", gr.grupo or "?"], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")
        for i, val in enumerate(gr.sk_answers):
            cell = ws.cell(row=r, column=i+4,
                           value=val if val is not None else "-")
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
        avg = ws.cell(row=r, column=14,
                      value=gr.sk_average if gr.sk_average else "-")
        avg.alignment = Alignment(horizontal="center")
        avg.font = Font(name="Arial", size=10, bold=True)

    for col, w in enumerate([8,8,8]+[7]*10+[10], 1):
        _cw(ws, col, w)


def _charts(wb, results):
    ws = wb.create_sheet("Gráficas")
    ws.sheet_view.showGridLines = False
    ws["A1"].value = "Gráficas de Resultados"
    ws["A1"].font  = Font(bold=True, size=14, name="Arial", color=C_HDR_BG)
    ws.row_dimensions[1].height = 24

    ds = 3
    for col, hdr in enumerate(["Folio","Porcentaje","Prom. SK",
                                "Grado","Grupo"], 1):
        ws.cell(row=ds, column=col, value=hdr)

    for i, gr in enumerate(results):
        row = ds + 1 + i
        ws.cell(row=row, column=1, value=gr.folio)
        ws.cell(row=row, column=2, value=gr.percentage)
        ws.cell(row=row, column=3, value=gr.sk_average or 0)
        ws.cell(row=row, column=4, value=gr.grado or "?")
        ws.cell(row=row, column=5, value=gr.grupo or "?")

    last = ds + len(results)

    bar = BarChart()
    bar.type  = "col"
    bar.title = "Puntaje por Estudiante (%)"
    bar.style = 10
    bar.y_axis.title = "Porcentaje"
    bar.x_axis.title = "Folio"
    bar.height = 12; bar.width = 20
    bar.y_axis.scaling.min = 0
    bar.y_axis.scaling.max = 100
    bar.add_data(Reference(ws, min_col=2, min_row=ds, max_row=last),
                 titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=ds+1, max_row=last))
    ws.add_chart(bar, "H3")

    line = LineChart()
    line.title = "Promedio Autoconocimiento"
    line.style = 10
    line.y_axis.title = "Promedio (1-5)"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 5
    line.height = 12; line.width = 20
    line.add_data(Reference(ws, min_col=3, min_row=ds, max_row=last),
                  titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=ds+1, max_row=last))
    ws.add_chart(line, "H23")

    for col, w in enumerate([10,14,14,10,10], 1):
        _cw(ws, col, w)


def _clave_sheet(wb, answer_key, exam_name):
    ws = wb.create_sheet("Clave de Respuestas")
    ws.sheet_view.showGridLines = False
    n        = len(answer_key)
    last_col = get_column_letter(n + 1)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].value     = f"Clave de Respuestas — {exam_name}"
    ws["A1"].font      = Font(bold=True, size=13, color=C_HDR_FG, name="Arial")
    ws["A1"].fill      = PatternFill("solid", fgColor=C_HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 3 — question numbers
    lbl = ws.cell(row=3, column=1, value="Pregunta")
    lbl.font = Font(bold=True, name="Arial", size=10)
    lbl.fill = PatternFill("solid", fgColor="ECF0F1")
    lbl.alignment = Alignment(horizontal="center")
    for i in range(n):
        cell = ws.cell(row=3, column=i + 2, value=i + 1)
        cell.font      = Font(bold=True, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor="ECF0F1")
        cell.alignment = Alignment(horizontal="center")

    # Row 4 — answers
    lbl2 = ws.cell(row=4, column=1, value="Respuesta")
    lbl2.font      = Font(bold=True, name="Arial", size=10, color=C_HDR_FG)
    lbl2.fill      = PatternFill("solid", fgColor=C_ACCENT)
    lbl2.alignment = Alignment(horizontal="center")
    for i, ans in enumerate(answer_key):
        cell = ws.cell(row=4, column=i + 2, value=ans)
        cell.font      = Font(bold=True, name="Arial", size=11)
        cell.fill      = PatternFill("solid", fgColor=C_GREEN)
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    for i in range(n):
        ws.column_dimensions[get_column_letter(i + 2)].width = 5


# ── Reader helpers ─────────────────────────────────────────────────────────────

def read_answer_key_from_excel(excel_path):
    """Return the answer key list from a previously exported Excel, or None."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, data_only=True)
    if "Clave de Respuestas" not in wb.sheetnames:
        return None
    ws = wb["Clave de Respuestas"]
    for row in ws.iter_rows():
        if row[0].value == "Respuesta":
            return [cell.value for cell in row[1:] if cell.value is not None]
    return None


def read_session_from_excel(excel_path):
    """
    Read a previously exported session.
    Returns (exam_name, answer_key, rows) where each row is a dict with:
        page_num, folio, grado, grupo, score, total, percentage,
        sk_average, confidence, error, mc_answers, sk_answers
    """
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, data_only=True)

    # Exam name
    exam_name = "Examen"
    if "Resumen" in wb.sheetnames:
        title = wb["Resumen"]["A1"].value or ""
        if "—" in str(title):
            exam_name = str(title).split("—", 1)[1].strip()

    answer_key = read_answer_key_from_excel(excel_path) or []
    n_q        = len(answer_key)

    # Resumen sheet → base student data
    resumen = {}
    if "Resumen" in wb.sheetnames:
        for row in wb["Resumen"].iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                break
            folio = str(row[0])
            resumen[folio] = {
                "folio":      folio,
                "grado":      str(row[1]) if row[1] else None,
                "grupo":      str(row[2]) if row[2] else None,
                "score":      row[3] or 0,
                "total":      row[4] or n_q,
                "percentage": row[5] or 0.0,
                "sk_average": row[7] if row[7] not in (None, "N/A") else None,
                "confidence": (row[8] or 0) / 100,
                "error":      str(row[9]) if row[9] and row[9] != "OK" else None,
            }

    # Detalle Preguntas → page_num + mc_answers
    detail = {}
    if "Detalle Preguntas" in wb.sheetnames:
        for row in wb["Detalle Preguntas"].iter_rows(min_row=4, values_only=True):
            if row[0] is None or row[1] is None:
                break
            folio = str(row[1])
            detail[folio] = {
                "page_num":   int(row[0]) if row[0] else 0,
                "mc_answers": [v if v != "-" else None
                               for v in row[4:4 + n_q]],
            }

    # Autoconocimiento → sk_answers
    sk_data = {}
    if "Autoconocimiento" in wb.sheetnames:
        for row in wb["Autoconocimiento"].iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                break
            folio = str(row[0])
            sk_data[folio] = [v if v != "-" else None for v in row[3:13]]

    # Merge and sort
    rows = []
    for folio, rd in resumen.items():
        dd  = detail.get(folio, {})
        row = {**rd,
               "page_num":   dd.get("page_num", 0),
               "mc_answers": dd.get("mc_answers", [None] * n_q),
               "sk_answers": sk_data.get(folio, [None] * 10)}
        rows.append(row)
    rows.sort(key=lambda r: r["page_num"])
    return exam_name, answer_key, rows
