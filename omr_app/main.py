"""
OMR Grader — Desktop GUI
"""
import os, sys, json, threading, tkinter as tk
from tkinter import ttk, filedialog, messagebox
import datetime

from sheet_generator import generate_omr_sheet
from omr_scanner     import scan_pdf, scan_single_page_from_pdf
from omr_grader      import grade_results, GradeResult
from omr_exporter    import export_to_excel, read_session_from_excel

CONFIG_FILE = os.path.join(os.path.dirname(__file__), "omr_config.json")

def load_config():
    d = {"exam_name": "Examen", "num_mc_questions": 20,
         "answer_key": [],
         "last_output_dir": os.path.expanduser("~")}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                d.update(json.load(f))
        except Exception:
            pass
    return d

def save_config(cfg):
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

# Day theme — inspired by Fundación Educa México logo
BG     = "#EEF2F7"   # light blue-gray page background
BG2    = "#FFFFFF"   # white inputs
ACCENT = "#2B3990"   # navy blue (logo primary)
SUCCESS= "#1A7A42"   # muted green
ERROR  = "#C0392B"   # red
FG     = "#1A202C"   # near-black text
FG2    = "#546E7A"   # slate gray secondary text
CARD   = "#FFFFFF"   # white cards
PURPLE = "#5E35B1"   # purple (load session)


class OMRApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.cfg = load_config()
        self.title("Calificador de Exámenes")
        self.geometry("940x740")
        self.configure(bg=BG)
        self.resizable(True, True)
        self._apply_theme()

        self.pdf_path          = tk.StringVar()
        self.exam_name_var     = tk.StringVar(value=self.cfg["exam_name"])
        self.num_questions_var = tk.IntVar(value=self.cfg["num_mc_questions"])
        self.status_var        = tk.StringVar(value="Listo.")
        self.progress_var      = tk.DoubleVar(value=0)
        self.rescan_page_var   = tk.IntVar(value=1)

        # Session state — populated after grading or loading from Excel
        self.session    = None   # dict: pdf_path, out_path, exam_name,
                                 #       answer_key, n_questions, graded
        self.student_db = {}     # folio str → nombre str

        self._build_ui()

    # ── Theme ─────────────────────────────────────────────────────────────────
    def _apply_theme(self):
        s = ttk.Style(self)
        s.theme_use("clam")

        s.configure("TNotebook",         background=BG,     borderwidth=0)
        s.configure("TNotebook.Tab",     background=BG2,    foreground=FG2,
                    padding=[10, 4],     font=("Arial", 9))
        s.map("TNotebook.Tab",
              background=[("selected", ACCENT)],
              foreground=[("selected", "white")])

        s.configure("Treeview",
                    background=BG2,     foreground=FG,
                    fieldbackground=BG2, rowheight=22,
                    borderwidth=0,       font=("Arial", 9))
        s.configure("Treeview.Heading",
                    background=ACCENT,  foreground="white",
                    font=("Arial", 9, "bold"), relief="flat")
        s.map("Treeview", background=[("selected", ACCENT)],
              foreground=[("selected", "white")])

        s.configure("TProgressbar", troughcolor=BG, background=ACCENT,
                    borderwidth=0)
        s.configure("TScrollbar",   background=BG2, troughcolor=BG,
                    borderwidth=0, arrowsize=12)

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        hdr = tk.Frame(self, bg="white")
        hdr.pack(fill="x")
        tk.Frame(self, bg=ACCENT, height=3).pack(fill="x")   # navy bottom border

        try:
            from PIL import Image, ImageTk
            _dir = os.path.dirname(__file__)
            h = 44

            # Right logo first — must be packed before the center fill widget
            abe_img = Image.open(os.path.join(_dir, "ABE-logo.png")).convert("RGBA")
            abe_img = abe_img.resize(
                (int(abe_img.width * h / abe_img.height), h), Image.LANCZOS)
            self._abe_photo = ImageTk.PhotoImage(abe_img)
            tk.Label(hdr, image=self._abe_photo, bg="white").pack(
                side="right", padx=(0, 12), pady=6)

            # Left logo
            educa_img = Image.open(os.path.join(_dir, "logo.png")).convert("RGBA")
            educa_img = educa_img.resize(
                (int(educa_img.width * h / educa_img.height), h), Image.LANCZOS)
            self._educa_photo = ImageTk.PhotoImage(educa_img)
            tk.Label(hdr, image=self._educa_photo, bg="white").pack(
                side="left", padx=(12, 0), pady=6)
        except Exception:
            pass

        # Title — centered in remaining space
        tk.Label(hdr, text="Calificador de Exámenes", bg="white", fg=ACCENT,
                 font=("Arial", 18, "bold")).pack(expand=True, pady=10)

        main = tk.Frame(self, bg=BG)
        main.pack(fill="both", expand=True, padx=18, pady=14)

        # ── Scrollable left panel ──────────────────────────────────────────────
        left_outer = tk.Frame(main, bg=BG)
        left_outer.pack(side="left", fill="both", expand=True, padx=(0, 10))

        left_canvas = tk.Canvas(left_outer, bg=BG, highlightthickness=0)
        left_sb     = ttk.Scrollbar(left_outer, orient="vertical",
                                    command=left_canvas.yview)
        left_canvas.configure(yscrollcommand=left_sb.set)
        left_sb.pack(side="right", fill="y")
        left_canvas.pack(side="left", fill="both", expand=True)

        left = tk.Frame(left_canvas, bg=BG)
        win  = left_canvas.create_window((0, 0), window=left, anchor="nw")

        def _on_frame_resize(e):
            left_canvas.configure(scrollregion=left_canvas.bbox("all"))
        def _on_canvas_resize(e):
            left_canvas.itemconfig(win, width=e.width)
        def _on_mousewheel(e):
            left_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

        left.bind("<Configure>", _on_frame_resize)
        left_canvas.bind("<Configure>", _on_canvas_resize)
        left_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # ── Right panel ────────────────────────────────────────────────────────
        right = tk.Frame(main, bg=BG, width=360)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        self._card(left, "Configuración del Examen", self._build_settings)
        self._card(left, "Archivo PDF Escaneado",    self._build_pdf_zone)
        self._card(left, "Lista de Alumnos",         self._build_student_db)
        self._card(left, "Clave de Respuestas",      self._build_answer_key)
        self._build_actions(left)
        self._build_progress(left)
        self._build_right_panel(right)

    def _card(self, parent, title, builder_fn):
        frame = tk.Frame(parent, bg=CARD)
        frame.pack(fill="x", pady=(0, 10))
        tk.Label(frame, text=title, bg=CARD, fg=ACCENT,
                 font=("Arial", 10, "bold")).pack(anchor="w", padx=12, pady=(10, 4))
        tk.Frame(frame, bg=ACCENT, height=1).pack(fill="x", padx=12)
        inner = tk.Frame(frame, bg=CARD)
        inner.pack(fill="x", padx=12, pady=8)
        builder_fn(inner)

    def _build_settings(self, f):
        r1 = tk.Frame(f, bg=CARD); r1.pack(fill="x", pady=2)
        tk.Label(r1, text="Nombre del examen:", bg=CARD, fg=FG,
                 font=("Arial", 9)).pack(side="left")
        tk.Entry(r1, textvariable=self.exam_name_var, bg=BG2, fg=FG,
                 insertbackground=FG, font=("Arial", 10), bd=0,
                 width=26).pack(side="left", padx=(8, 0))

        r2 = tk.Frame(f, bg=CARD); r2.pack(fill="x", pady=6)
        tk.Label(r2, text="Nº preguntas Sec.1:", bg=CARD, fg=FG,
                 font=("Arial", 9)).pack(side="left")
        self.spinbox = tk.Spinbox(r2, from_=5, to=40,
                                  textvariable=self.num_questions_var,
                                  width=4, bg=BG2, fg=FG,
                                  buttonbackground=BG2, font=("Arial", 10),
                                  command=self._rebuild_answer_key)
        self.spinbox.pack(side="left", padx=(6, 0))
        self.spinbox.bind("<FocusOut>", lambda e: self._safe_rebuild())
        self.spinbox.bind("<Return>",   lambda e: self._safe_rebuild())

        r3 = tk.Frame(f, bg=CARD); r3.pack(fill="x", pady=(6, 0))
        self._btn(r3, "Generar Hoja OMR", self._generate_sheet, ACCENT)

    def _safe_rebuild(self):
        try:
            n = self.num_questions_var.get()
            if 1 <= n <= 40:
                self._rebuild_answer_key()
        except Exception:
            pass

    def _build_pdf_zone(self, f):
        self.drop_frame = tk.Frame(f, bg=BG2, bd=2, relief="solid", cursor="hand2")
        self.drop_frame.pack(fill="x", ipady=14)
        self.drop_label = tk.Label(self.drop_frame,
            text="Haz clic para seleccionar el PDF escaneado",
            bg=BG2, fg=FG2, font=("Arial", 10))
        self.drop_label.pack(pady=10)
        self.drop_frame.bind("<Button-1>", lambda e: self._browse_pdf())
        self.drop_label.bind("<Button-1>", lambda e: self._browse_pdf())

    def _build_student_db(self, f):
        row = tk.Frame(f, bg=CARD); row.pack(fill="x", pady=2)
        self._student_db_label = tk.Label(
            row, text="Sin lista cargada", bg=CARD, fg=FG2, font=("Arial", 9))
        self._student_db_label.pack(side="left", expand=True, anchor="w")
        self._btn(row, "Cargar Excel", self._browse_student_db, ACCENT)

    def _browse_student_db(self):
        path = filedialog.askopenfilename(
            title="Seleccionar lista de alumnos (Excel)",
            filetypes=[("Excel", "*.xlsx *.xls")])
        if not path:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            db = {}
            for row in ws.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                folio_val  = row[0]
                nombre_val = row[1] if len(row) > 1 else None
                if nombre_val is None:
                    continue
                if isinstance(folio_val, (int, float)):
                    folio_str = str(int(folio_val))
                else:
                    folio_str = str(folio_val).strip()
                if folio_str.lower() in ("folio", ""):
                    continue  # skip header row
                db[folio_str] = str(nombre_val).strip()
            self.student_db = db
            self._student_db_label.configure(
                text=f"{len(db)} alumno{'s' if len(db) != 1 else ''} cargados",
                fg=SUCCESS)
            self._log(f"Lista de alumnos: {len(db)} registros cargados")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer la lista:\n{e}")

    def _build_answer_key(self, f):
        ctrl = tk.Frame(f, bg=CARD)
        ctrl.pack(fill="x", pady=(0, 6))
        reset = tk.Label(ctrl, text="Resetear a A", bg=ERROR, fg="white",
                         font=("Arial", 8, "bold"), cursor="hand2",
                         padx=8, pady=3)
        reset.pack(side="right")
        reset.bind("<Button-1>", lambda e: self._reset_answer_key())

        self.key_frame = tk.Frame(f, bg=CARD)
        self.key_frame.pack(fill="x")
        self._rebuild_answer_key()

    def _rebuild_answer_key(self):
        for w in self.key_frame.winfo_children():
            w.destroy()
        n = self.num_questions_var.get()
        self.answer_vars = []
        self._key_btns  = []
        cols = 5

        for i in range(n):
            row = i // cols; col = i % cols
            sub = tk.Frame(self.key_frame, bg=CARD)
            sub.grid(row=row, column=col, padx=4, pady=3)
            tk.Label(sub, text=f"{i+1}", bg=CARD, fg=FG2,
                     font=("Arial", 7)).pack()

            var = tk.StringVar(value="A")
            if i < len(self.cfg.get("answer_key", [])):
                var.set(self.cfg["answer_key"][i])

            btn_row = tk.Frame(sub, bg=CARD)
            btn_row.pack()
            btns = {}
            for ch in ["A", "B", "C", "D"]:
                b = tk.Label(btn_row, text=ch, width=2,
                             font=("Arial", 8, "bold"), cursor="hand2",
                             padx=2, pady=2, relief="flat")
                b.pack(side="left", padx=1)
                btns[ch] = b

            def _make_select(v, bd, choice):
                def select(e=None):
                    v.set(choice)
                    for c, b in bd.items():
                        b.configure(bg=ACCENT if c == choice else BG2,
                                    fg="white" if c == choice else FG2)
                return select

            current = var.get()
            for ch, b in btns.items():
                b.configure(bg=ACCENT if ch == current else BG2,
                            fg="white" if ch == current else FG2)
                b.bind("<Button-1>", _make_select(var, btns, ch))

            self.answer_vars.append(var)
            self._key_btns.append(btns)

    def _reset_answer_key(self):
        for var in self.answer_vars:
            var.set("A")
        for btns in self._key_btns:
            for ch, b in btns.items():
                b.configure(bg=ACCENT if ch == "A" else BG2,
                            fg="white" if ch == "A" else FG2)

    def _build_actions(self, f):
        bf = tk.Frame(f, bg=BG); bf.pack(fill="x", pady=4)
        self._btn(bf, "Calificar PDF",        self._start_grading, SUCCESS)
        self._btn(bf, "Cargar sesión previa", self._load_session,  PURPLE)

    def _btn(self, parent, text, cmd, color):
        """Frame+Label button — macOS respects bg/fg on these unlike tk.Button."""
        f = tk.Frame(parent, bg=color, cursor="hand2")
        f.pack(side="left", padx=(0, 8))
        lbl = tk.Label(f, text=text, bg=color, fg="white",
                       font=("Arial", 10, "bold"), cursor="hand2",
                       padx=14, pady=8)
        lbl.pack()
        for w in (f, lbl):
            w.bind("<Button-1>", lambda e, c=cmd: c())

    def _build_progress(self, f):
        pf = tk.Frame(f, bg=BG); pf.pack(fill="x", pady=(6, 0))
        ttk.Progressbar(pf, variable=self.progress_var,
                        maximum=100).pack(fill="x", pady=(0, 4))
        tk.Label(pf, textvariable=self.status_var, bg=BG, fg=FG2,
                 font=("Arial", 9)).pack(anchor="w")

    # ── Right panel — tabbed ───────────────────────────────────────────────────
    def _build_right_panel(self, f):
        nb = ttk.Notebook(f)
        nb.pack(fill="both", expand=True)
        self.right_nb = nb

        log_tab = tk.Frame(nb, bg=BG)
        nb.add(log_tab, text="Registro")
        self._build_log_tab(log_tab)

        res_tab = tk.Frame(nb, bg=BG)
        nb.add(res_tab, text="Resultados")
        self._build_results_tab(res_tab)

    def _build_log_tab(self, f):
        tk.Label(f, text="Registro", bg=BG, fg=ACCENT,
                 font=("Arial", 11, "bold")).pack(anchor="w", padx=10, pady=(6, 4))
        self.log_box = tk.Text(f, bg=BG, fg=FG, font=("Courier", 8),
                               bd=0, state="disabled", wrap="word")
        sb = ttk.Scrollbar(f, command=self.log_box.yview)
        sb.pack(fill="y", side="right")
        self.log_box.pack(fill="both", expand=True, padx=(8, 0))
        self.log_box.configure(yscrollcommand=sb.set)
        tk.Label(f, text="─" * 44, bg=BG, fg=BG2).pack()
        self._btn(f, "Limpiar", self._clear_log, FG2)

    def _build_results_tab(self, f):
        tk.Label(f, text="Resultados", bg=BG, fg=ACCENT,
                 font=("Arial", 11, "bold")).pack(anchor="w", padx=10, pady=(6, 4))

        # Treeview
        cols = ("page", "folio", "nombre", "grado", "grupo", "score", "pct", "status")
        tf   = tk.Frame(f, bg=BG)
        tf.pack(fill="both", expand=True, padx=4)

        self.results_tree = ttk.Treeview(tf, columns=cols, show="headings",
                                         height=22, selectmode="browse")
        col_cfg = [("Pág.", 32), ("Folio", 38), ("Nombre", 80), ("Gdo.", 28),
                   ("Grp.", 28), ("Punt.", 36), ("%", 34), ("Estado", 46)]
        for col, (hdr, w) in zip(cols, col_cfg):
            self.results_tree.heading(col, text=hdr)
            self.results_tree.column(col, width=w, anchor="center", stretch=False)

        vsb = ttk.Scrollbar(tf, orient="vertical",
                            command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=vsb.set)
        self.results_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.results_tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        # Re-scan controls
        rc = tk.Frame(f, bg=BG); rc.pack(fill="x", padx=6, pady=(6, 2))
        tk.Label(rc, text="Pág. PDF:", bg=BG, fg=FG2,
                 font=("Arial", 9)).pack(side="left")
        tk.Spinbox(rc, from_=1, to=999, textvariable=self.rescan_page_var,
                   width=4, bg=BG2, fg=FG, font=("Arial", 9),
                   buttonbackground=BG2).pack(side="left", padx=4)
        self._btn(rc, "Re-scan", self._rescan_page, ACCENT)

        tk.Frame(f, bg=BG, height=4).pack()
        tk.Label(f, text="sin sesión activa", bg=BG, fg=FG2,
                 font=("Arial", 8, "italic")).pack()

    # ── helpers ───────────────────────────────────────────────────────────────
    def _log(self, msg):
        self.log_box.configure(state="normal")
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_box.insert("end", f"[{ts}] {msg}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="Seleccionar PDF escaneado",
            filetypes=[("PDF", "*.pdf")])
        if path:
            self.pdf_path.set(path)
            self.drop_label.configure(
                text=f"  {os.path.basename(path)}", fg=SUCCESS)
            self._log(f"PDF cargado: {os.path.basename(path)}")

    def _get_answer_key(self):
        return [v.get() for v in self.answer_vars]

    def _read_n(self):
        try:
            n = int(self.spinbox.get())
            self.num_questions_var.set(n)
            return n
        except Exception:
            return self.num_questions_var.get()

    # ── results table ─────────────────────────────────────────────────────────
    def _populate_results_table(self):
        for row in self.results_tree.get_children():
            self.results_tree.delete(row)
        if not self.session:
            return
        for gr in self.session["graded"]:
            status = "ERROR" if gr.error else ("OK" if gr.percentage >= 70 else "")
            tag    = "err" if gr.error else ("ok" if gr.percentage >= 70 else "low")
            nombre = self.student_db.get(str(gr.folio), "")
            self.results_tree.insert("", "end", iid=str(gr.page_num),
                values=(gr.page_num, gr.folio, nombre, gr.grado or "?",
                        gr.grupo or "?", f"{gr.score}/{gr.total}",
                        f"{gr.percentage}%", status),
                tags=(tag,))
        self.results_tree.tag_configure("ok",  background="#D4EDDA", foreground="#155724")
        self.results_tree.tag_configure("low", background="#F8D7DA", foreground="#721C24")
        self.results_tree.tag_configure("err", background="#FFF3CD", foreground="#856404")
        self.right_nb.select(1)   # switch to Resultados tab

    def _on_tree_select(self, event):
        sel = self.results_tree.selection()
        if sel:
            self.rescan_page_var.set(int(sel[0]))

    # ── generate sheet ────────────────────────────────────────────────────────
    def _generate_sheet(self):
        n    = self._read_n()
        exam = self.exam_name_var.get() or "Examen"
        odir = self.cfg.get("last_output_dir", os.path.expanduser("~"))
        path = filedialog.asksaveasfilename(
            title="Guardar hoja OMR", defaultextension=".pdf",
            initialdir=odir, filetypes=[("PDF", "*.pdf")],
            initialfile=f"hoja_omr_{exam.replace(' ', '_')}.pdf")
        if not path:
            return
        try:
            generate_omr_sheet(path, num_mc_questions=n, exam_name=exam)
            self._log(f"Hoja OMR generada: {os.path.basename(path)}")
            messagebox.showinfo("Éxito", f"Hoja OMR guardada en:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ── grading ───────────────────────────────────────────────────────────────
    def _start_grading(self):
        if not self.pdf_path.get():
            messagebox.showwarning("Aviso", "Selecciona un PDF escaneado primero.")
            return
        answer_key = self._get_answer_key()
        exam       = self.exam_name_var.get() or "Examen"
        n          = self._read_n()
        odir       = self.cfg.get("last_output_dir", os.path.expanduser("~"))
        out        = filedialog.asksaveasfilename(
            title="Guardar resultados Excel", defaultextension=".xlsx",
            initialdir=odir, filetypes=[("Excel", "*.xlsx")],
            initialfile=f"resultados_{exam.replace(' ', '_')}.xlsx")
        if not out:
            return

        self.cfg.update({"last_output_dir": os.path.dirname(out),
                         "exam_name": exam, "num_mc_questions": n,
                         "answer_key": answer_key})
        save_config(self.cfg)

        threading.Thread(
            target=self._run_grading,
            args=(self.pdf_path.get(), answer_key, exam, out, n),
            daemon=True).start()

    def _run_grading(self, pdf_path, answer_key, exam_name, out_path, n_questions):
        try:
            self.status_var.set("Procesando PDF...")
            self.progress_var.set(5)
            self._log(f"Iniciando escaneo: {os.path.basename(pdf_path)}")

            def cb(i, total):
                self.progress_var.set(10 + int(i / max(total, 1) * 70))
                self.status_var.set(f"Procesando página {i+1}/{total}...")
                self._log(f"  Página {i+1}/{total}")

            scans  = scan_pdf(pdf_path, n_questions, progress_callback=cb)
            self._log(f"Páginas procesadas: {len(scans)}")
            self.progress_var.set(80)

            self.status_var.set("Calificando...")
            graded = grade_results(scans, answer_key)
            for gr in graded:
                s = "✓" if gr.percentage >= 70 else "✗"
                self._log(f"  {s} Folio {gr.folio} | "
                          f"{gr.score}/{gr.total} ({gr.percentage}%)")

            self.progress_var.set(90)
            self.status_var.set("Exportando Excel...")
            export_to_excel(graded, answer_key, exam_name, out_path,
                            student_db=self.student_db or None)

            # Store session for re-scan
            self.session = {"pdf_path":    pdf_path,
                            "out_path":    out_path,
                            "exam_name":   exam_name,
                            "answer_key":  answer_key,
                            "n_questions": n_questions,
                            "graded":      graded}

            self.progress_var.set(100)
            self.status_var.set(f"Listo — {out_path}")
            self._log(f"Excel guardado: {os.path.basename(out_path)}")
            self.after(0, self._populate_results_table)
            self.after(0, lambda: messagebox.showinfo(
                "Calificación Completa",
                f"Se calificaron {len(graded)} estudiantes.\n\n"
                f"Archivo guardado en:\n{out_path}"))
        except Exception as e:
            self.status_var.set(f"Error: {e}")
            self._log(f"ERROR: {e}")
            self.after(0, lambda: messagebox.showerror("Error", str(e)))

    # ── re-scan ───────────────────────────────────────────────────────────────
    def _rescan_page(self):
        if not self.session:
            messagebox.showwarning("Aviso",
                "No hay sesión activa. Califica un PDF o carga una sesión previa.")
            return
        page_num = self.rescan_page_var.get()
        threading.Thread(target=self._run_rescan, args=(page_num,),
                         daemon=True).start()

    def _run_rescan(self, page_num):
        try:
            s = self.session
            self.status_var.set(f"Re-escaneando página {page_num}...")
            self._log(f"Re-scan iniciado — pág. {page_num}")

            new_scan   = scan_single_page_from_pdf(s["pdf_path"], page_num,
                                                   s["n_questions"])
            new_graded = grade_results([new_scan], s["answer_key"])[0]

            # Replace matching entry or append
            graded = s["graded"]
            for i, gr in enumerate(graded):
                if gr.page_num == page_num:
                    graded[i] = new_graded
                    break
            else:
                graded.append(new_graded)
                graded.sort(key=lambda g: g.page_num)

            export_to_excel(graded, s["answer_key"], s["exam_name"], s["out_path"],
                            student_db=self.student_db or None)

            self.status_var.set(
                f"Re-scan completo — Folio {new_graded.folio} "
                f"({new_graded.score}/{new_graded.total})")
            self._log(f"  Re-scan pág.{page_num}: Folio {new_graded.folio} | "
                      f"{new_graded.score}/{new_graded.total} "
                      f"({new_graded.percentage}%)")
            self.after(0, self._populate_results_table)
        except Exception as e:
            self.status_var.set(f"Error: {e}")
            self._log(f"ERROR re-scan: {e}")
            self.after(0, lambda: messagebox.showerror("Error re-scan", str(e)))

    # ── load session from Excel ───────────────────────────────────────────────
    def _load_session(self):
        excel_path = filedialog.askopenfilename(
            title="Seleccionar Excel de resultados previos",
            filetypes=[("Excel", "*.xlsx")])
        if not excel_path:
            return

        pdf_path = filedialog.askopenfilename(
            title="Seleccionar el PDF original escaneado",
            filetypes=[("PDF", "*.pdf")])
        if not pdf_path:
            return

        try:
            exam_name, answer_key, rows = read_session_from_excel(excel_path)
            if not answer_key:
                messagebox.showerror("Error",
                    "No se encontró la hoja 'Clave de Respuestas' en el Excel.\n"
                    "Solo se pueden cargar archivos generados por esta aplicación.")
                return

            n_q    = len(answer_key)
            graded = []
            for row in rows:
                mc_correct = [
                    (ans == answer_key[i]) if ans is not None and i < n_q else False
                    for i, ans in enumerate(row["mc_answers"])
                ]
                sk    = [v for v in row["sk_answers"] if v is not None]
                sk_avg= round(sum(sk) / len(sk), 2) if sk else None
                graded.append(GradeResult(
                    page_num   = row["page_num"],
                    folio      = row["folio"],
                    grado      = row["grado"],
                    grupo      = row["grupo"],
                    mc_answers = row["mc_answers"],
                    mc_correct = mc_correct,
                    score      = row["score"],
                    total      = row["total"],
                    percentage = row["percentage"],
                    sk_answers = row["sk_answers"],
                    sk_average = sk_avg,
                    confidence = row["confidence"],
                    error      = row.get("error"),
                ))

            self.session = {"pdf_path":    pdf_path,
                            "out_path":    excel_path,
                            "exam_name":   exam_name,
                            "answer_key":  answer_key,
                            "n_questions": n_q,
                            "graded":      graded}

            self._log(f"Sesión cargada: {os.path.basename(excel_path)} "
                      f"({len(graded)} estudiantes)")
            self.status_var.set(f"Sesión cargada — {len(graded)} estudiantes")
            self._populate_results_table()

        except Exception as e:
            messagebox.showerror("Error al cargar sesión", str(e))
            self._log(f"ERROR cargar sesión: {e}")


if __name__ == "__main__":
    OMRApp().mainloop()
